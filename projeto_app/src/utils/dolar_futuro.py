from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from networkday import networkdays


try: 
    base_frc = load_workbook(r"C:\Users\YuriSaneripCalzzani\Documents\Python\FRC.xlsx")
    base_di1 = load_workbook(r"C:\Users\YuriSaneripCalzzani\Documents\Python\DI1.xlsx")
except Exception as e:
    print(f'Erro ao carregar planilha: {e}')


def calcular_fator_di1(lista_di1: list, lista_contratos: list):
    print(lista_di1)
    fator_di1 =[]

    mes_string_para_numerico =  {
        'Jan' : '01-01',
        'Fev': '02-01',
        'Mar': '03-01',
        'Abr': '04-01',
        'Mai' : '05-01',
        'Jun' : '06-01',
        'Jul': '07-01',
        'Ago' : '08-01',
        'Set' : '09-01',
        'Out' : '10-01',
        'Nov' : '11-01',
        'Dez' : '12-01'
    }

    contador = 0
    for contrato in lista_contratos:
        data_vencimento = f'{contrato[-4:]}-{mes_string_para_numerico[contrato[:3]]}'
        dias_uteis = networkdays(str(datetime.now().strftime('%Y-%m-%d')), str(datetime.strptime(data_vencimento, '%Y-%m-%d')))[0]
        print(lista_di1[contador])
        if lista_di1[contador] != 0:
            fator_di1.append(((1 + lista_di1[contador]/100)**(dias_uteis/252)).astype(float))
        else:
            fator_di1.append(0)
        contador +=1
    
    return fator_di1

def calcular_fator_frc(lista_frc: list, lista_contratos: list):
    
    fator_frc =[]

    mes_string_para_numerico =  {
        'Jan' : '01-01',
        'Fev': '02-01',
        'Mar': '03-01',
        'Abr': '04-01',
        'Mai' : '05-01',
        'Jun' : '06-01',
        'Jul': '07-01',
        'Ago' : '08-01',
        'Set' : '09-01',
        'Out' : '10-01',
        'Nov' : '11-01',
        'Dez' : '12-01'
    }

    contador = 0
    for contrato in lista_contratos:
        data_vencimento = f'{contrato[-4:]}-{mes_string_para_numerico[contrato[:3]]}'
        dias_corridos = abs((datetime.strptime(data_vencimento, '%Y-%m-%d').date()- date.today()).days)
        print(dias_corridos)
        if lista_frc[contador] != 0:
            fator_frc.append((lista_frc[contador]/100)*(dias_corridos/360)+1)
        else:
            fator_frc.append(0)
        contador +=1
    
    return fator_frc


def extrair_contratos():

    mes_numerico_para_string =  {
        '1': 'Jan',
        '2': 'Fev',
        '3': 'Mar',
        '4': 'Abr',
        '5': 'Mai',
        '6': 'Jun',
        '7': 'Jul',
        '8': 'Ago',
        '9': 'Set',
        '10': 'Out',
        '11': 'Nov',
        '12': 'Dez'
    }

    mes = int(datetime.now().strftime('%m'))
    ano = int(datetime.now().strftime('%Y'))

    contratos = []
    if mes > 10:
        ano += 1
    ano_conclusao = ano + 3

    if mes <= 3 or mes > 10:
        mes = 3
        while ano <= ano_conclusao:
            contratos.append(f'{mes_numerico_para_string[str(mes)]}-{ano}')
            if mes == 7:
                mes += 3
            elif mes == 10:
                mes = (mes - 10) + 3
                ano += 1
            else:
                mes += 2
        
    elif mes > 3 and mes <= 5:
        mes = 5
        while ano <= ano_conclusao:
            contratos.append(f'{mes_numerico_para_string[str(mes)]}-{ano}')
            if mes == 7:
                mes += 3
            elif mes == 10:
                mes = (mes - 10) + 3
                ano += 1
            else:
                mes += 2

    elif mes > 5 and mes <= 7:
        mes = 7
        while ano <= ano_conclusao:
            contratos.append(f'{mes_numerico_para_string[str(mes)]}-{ano}')
            if mes == 7:
                mes += 3
            elif mes == 10:
                mes = (mes - 10) + 3
                ano += 1
            else:
                mes += 2

    elif mes > 7 and mes <= 10:
        mes = 10
        while ano <= ano_conclusao:
            contratos.append(f'{mes_numerico_para_string[str(mes)]}-{ano}')
            if mes == 7:
                mes += 3
            elif mes == 10:
                mes = (mes - 10) + 3
                ano += 1
            else:
                mes += 2
                
    contratos = contratos[:10]

    return contratos
    
def extrair_dados_planilha_frc(contratos: list):
    dados_frc = []
    
    for item in contratos:
        try :
            dados_frc.append(base_frc[item].cell(base_frc[item].max_row, 6).internal_value)
            
        except:
            dados_frc.append(0)
    
    return dados_frc
            

def extrair_dados_planilha_di1(contratos: list):
    dados_di1 = []
    
    for item in contratos:
        try :
            dados_di1.append(base_di1[item].cell(base_di1[item].max_row, 7).internal_value)
        except:
            dados_di1.append(0)

    return dados_di1
        

def calcula_dolar_futuro_para_açucar(dolar_spot: float, fator_di1: list, fator_frc : list, contratos : list):
    lista_dolar_futuro = []
    lista_contratos_com_valor = []
    contador = 0
    for item in fator_di1:
        lista_dolar_futuro.append(dolar_spot * (item/fator_frc[contador]))


    lista_de_indices = [index for index, value in enumerate(lista_dolar_futuro) if value == 0]
    lista_contratos_sem_valor = [contratos[indice] for indice in lista_de_indices]


    for item in lista_contratos_sem_valor:
        posterior = contratos.index(item) + 1
        anterior = contratos.index(item) - 1
        # if contratos[posterior] != 0:
        #     lista_contratos_com_valor.append(posterior)
        if contratos[anterior] != 0:
            lista_contratos_com_valor.append(anterior)
   

    print(lista_contratos_sem_valor)
    print('---------------------------------------------------------')
    print(lista_contratos_com_valor)
    



    return lista_dolar_futuro
    
    

contratos = extrair_contratos()
print(contratos)
frc = extrair_dados_planilha_frc(contratos)
di1 = extrair_dados_planilha_di1(contratos)
print(extrair_dados_planilha_frc(contratos))
resultado_1 = calcular_fator_di1(di1 ,contratos)
print(resultado_1)
resultado_2 = calcular_fator_frc(di1 ,contratos)
print(resultado_2)

print(calcula_dolar_futuro_para_açucar(6.14, resultado_1, resultado_2, contratos))


