from datetime import datetime
import datetime
import pandas as pd
import openpyxl
import asyncio
import sys 
sys.path.append(r"C:\Users\YuriSaneripCalzzani\Documents\Python")
from futuro_scrapers import extrair_dados_bmf_b3_para_json
from contrato_para_data_vencimento import contrato_para_data_vencimento
import json



#  datelist = pd.date_range(datetime.today().date(), origin=pd.Timestamp("2023-01-01"), periods=365).tolist()
index = pd.date_range(datetime.datetime(2024, 11, 25), datetime.datetime.now())
# try:
#     base = pd.ExcelFile(r"C:\Users\YuriSaneripCalzzani\Documents\Python\DI1-historico.xlsx").sheet_names
# except:
#     sheet_names = []


# for data_cotacao in index:
#     dados_bmf = asyncio.run(extrair_dados_bmf_b3_para_json('DI1', data_cotacao))
#     if dados_bmf != []:
#         contratos = []
#         contrato_dicionario = {}

#         for index in range(0, len(dados_bmf)):
#             contratos.append(dados_bmf[index]['VENCTO \xa0'])
#             del dados_bmf[index]['VENCTO \xa0']
#             contrato_dicionario[contratos[index]] = dados_bmf[index]
#         print(contrato_dicionario)

#         for contrato, dados_bmf in contrato_dicionario.items(): 
#             base = pd.read_excel(r"C:\Users\YuriSaneripCalzzani\Documents\Python\DI1-historico.xlsx")
#             new_row = pd.DataFrame([dados_bmf])

#             print(base)

#             base = pd.concat(
#                 [base, new_row], 
#                 ignore_index=True
#                 )
#             print(list(contrato_dicionario.keys()))
#             print(base)

#         if contrato != list(contrato_dicionario.keys()):
#             with pd.ExcelWriter(r"C:\Users\YuriSaneripCalzzani\Documents\Python\DI1-historico.xlsx",
#                 mode = 'a',
#                 if_sheet_exists='new', 
#                 engine = 'openpyxl') as writer:
#                 base.to_excel(writer, sheet_name=f'{contrato}', index=False)
#                 sheet_names = pd.ExcelFile(r"C:\Users\YuriSaneripCalzzani\Documents\Python\DI1-historico.xlsx").sheet_names
#         else: 
#             with pd.ExcelWriter(r"C:\Users\YuriSaneripCalzzani\Documents\Python\DI1-historico.xlsx",
#                 mode = 'a',
#                 if_sheet_exists='overlay', 
#                 engine = 'openpyxl') as writer:
#                 base.to_excel(writer, sheet_name=f'{contrato}', index=False)
        
import pandas as pd
from openpyxl import load_workbook

# Caminho do arquivo Excel
arquivo_excel = r"C:\Users\YuriSaneripCalzzani\Documents\Python\FRC.xlsx"

for data_cotacao in index:
    # Extrai os dados_bmf
    dados_bmf = asyncio.run(extrair_dados_bmf_b3_para_json('FRC', data_cotacao))
    

    
    for item in dados_bmf:
       item = contrato_para_data_vencimento(item)
       
    
    

    if dados_bmf:  # Se houver dados_bmf
        contratos = []
        contrato_dicionario = {}

        for index in range(len(dados_bmf)):
            contratos.append(dados_bmf[index]['VENCTO \xa0'])
            del dados_bmf[index]['VENCTO \xa0']
            contrato_dicionario[contratos[index]] = dados_bmf[index]

        # Carrega as planilhas existentes no arquivo
        try:
            workbook = load_workbook(arquivo_excel)
            planilhas_existentes = workbook.sheetnames
        except FileNotFoundError:
            planilhas_existentes = []

        for contrato, dados_bmf in contrato_dicionario.items():
            # Converte os dados_bmf em DataFrame
            new_row = pd.DataFrame([dados_bmf])

            # Verifica se a planilha já existe
            if contrato in planilhas_existentes:
                # Lê os dados_bmf existentes na planilha
                base = pd.read_excel(arquivo_excel, sheet_name=contrato)

                # Concatena os novos dados_bmf
                base = pd.concat([base, new_row], ignore_index=True)
            else:
                # Se a planilha não existir, cria um novo DataFrame
                base = new_row

            # Salva o DataFrame na planilha correspondente
            with pd.ExcelWriter(
                arquivo_excel, mode="a", if_sheet_exists="replace", engine="openpyxl"
            ) as writer:
                base.to_excel(writer, sheet_name=contrato, index=False)

            print(f"Dados salvos na planilha: {contrato}")


